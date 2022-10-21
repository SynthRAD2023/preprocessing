import pydicom as dcm
import os
import csv
import argparse
import openpyxl
import SimpleITK as sitk

# Function that reads a list of dicom tag names from a txt file and returns the tags in a list
def read_tags(txt_file):
    file = open(txt_file,'r')
    tags = file.read().splitlines()
    return tags

# Function to get the number of slices for a dicom image
def get_number_of_slices(dcm_folder):
    files = os.listdir(dcm_folder)
    slices = 0
    UID = []
    for i in range(len(files)):
        tags = dcm.dcmread(os.path.join(dcm_folder,files[i]),stop_before_pixels=True)
        UID.append(tags['SeriesInstanceUID'].value)
        if i==0:
            slices = 1
        elif i!=0:
            if UID[i]==UID[i-1]:
                slices = slices+1
    return slices

# Function that takes a DICOM folder as an input, reads dicom tags from the first slice and returns specified tags as a dict,
# and if specified also adds dimension and spacing of post-processed image
def extract_tags(dcm_folder,tags,pre_processed=None,csv=None,pt=None,phase=None):
    files = [filename for filename in os.listdir(dcm_folder) if (filename.startswith("ct")|filename.startswith("mr")) ]
#    files = os.listdir(dcm_folder)
    tag_list = read_tags(tags)
    tags = dcm.dcmread(os.path.join(dcm_folder,files[0]),stop_before_pixels=True)
    tags_dict = {}
    for tag in tag_list:
        try:
            tags_dict[tag]=str(tags[tag].value)
        except:
            print('Tag: ' + str(tag) + ' not available!')
            tags_dict[tag]='empty'
    tags_dict['Slices']=str(get_number_of_slices(dcm_folder))
    if pre_processed!=None:
        size_pre,spacing_pre = extract_tags_post(pre_processed)
        tags_dict['Dim_pre']=str(size_pre)
        tags_dict['Spacing_pre']=str(spacing_pre)
    if csv!=None:
        if pt != None:
            write_dict_to_csv(tags_dict, csv, tag_list, pt, phase)
        else:
            write_dict_to_csv(tags_dict, csv, tag_list)
   # print(tags_dict)
    return tags_dict

def extract_tags_post(image):
    im = sitk.ReadImage(image)
    imsize = im.GetSize()
    imspacing = im.GetSpacing()
    return [imsize,imspacing]

# Function that creates a csv file based on the dicts with extracted tags
def write_dict_to_csv(input_dict,output_csv,tag_list,pt,phase):
    
    # first check if dict is nested (required for file writing)

    if pt != None:
        input_dict['ID'] = pt
        input_dict['Set'] = phase
        tag_list.insert(0,'ID')
        tag_list.insert(1,'Set')

    if any(isinstance(i,dict) for i in input_dict.values()):
        input_dict_nested = input_dict
    else:
        input_dict_nested = {'1': input_dict}
#    print(input_dict_nested)
    tag_list.append('Slices')
    if 'Dim_pre' in input_dict:
        tag_list.append('Dim_pre')
        tag_list.append('Spacing_pre')

# If file does not exist, create header, otherwise directly append the rows
    if os.path.isfile(output_csv):
        print("File exist")
    else:
        with open(output_csv, 'w+', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=tag_list)
            writer.writeheader()

    with open(output_csv,'a', newline='') as csvfile:
        writer = csv.DictWriter(csvfile,fieldnames=tag_list)
        for k in input_dict_nested:
            writer.writerow({field: input_dict_nested[k].get(field) or k for field in tag_list})
        print('csv '+ output_csv +' written!')

#Function that converts the csv file into an excel file
def convert_csv_to_xlsx(input_csv, output_xlsx, sheetname=None):
    csv_data = []
    with open(input_csv) as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            csv_data.append(row)
    if os.path.isfile(output_xlsx):
        workbook = openpyxl.load_workbook(output_xlsx)
        std = workbook['Sheet']
        workbook.remove(std)
    else:
        workbook = openpyxl.Workbook()
    workbook.create_sheet(sheetname)
    sheet = workbook[sheetname]
    if sheetname!=None:
        sheet.title=sheetname
    for row in csv_data:
        sheet.append(row)
    workbook.save(output_xlsx)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Define fixed, moving and output filenames')
    parser.add_argument('operation', help='select operation to perform (extract, to_csv, to_xlsx)')
    parser.add_argument('--tags', help='dictionary file listing the tags to extract')
    parser.add_argument('--pre', help='path to preprocessed nii.gz to extract sizes')
    parser.add_argument('--path', help='path of the folder containing the dicom')
    parser.add_argument('--csv', help='path of the output csv file')
    parser.add_argument('--xlsx', help='path of the output excel file')
    parser.add_argument('--pt', help='ID of the patient')
    parser.add_argument('--phase', help='Dataset of the patient')
    args = parser.parse_args()

    if args.operation == 'extract':
        if args.pre is None:
            extract_tags(args.path,args.tags,pre_processed=None)
        else:
            if args.csv is None:
                extract_tags(args.path,args.tags,args.pre,csv=None)
            else:
                if args.pt is None:
                    extract_tags(args.path, args.tags, args.pre, args.csv, pt=None, phase=None)
                else:
                    extract_tags(args.path, args.tags, args.pre, args.csv, args.pt, args.phase)
    elif args.operation == 'toxlsx':
        convert_csv_to_xlsx(args.csv, args.xlsx, args.tags)
    else:
        print('check help for usage instructions')
